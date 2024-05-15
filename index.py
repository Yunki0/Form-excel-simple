# Import des modules nécessaires
import openpyxl
from kivy.app import App
from kivy.uix.gridlayout import GridLayout
from kivy.uix.label import Label
from kivy.uix.textinput import TextInput
from kivy.uix.button import Button
from kivy.uix.popup import Popup

# Définition de l'application Kivy
class FormulaireApp(App):
    def build(self):
        # Création de la mise en page de la grille avec deux colonnes
        self.layout = GridLayout(cols=2, padding=10, spacing=10)
        
        # Ajout des étiquettes et des champs de saisie de texte pour la désignation et la quantité
        self.layout.add_widget(Label(text="Designation:"))
        self.designation_input = TextInput()
        self.layout.add_widget(self.designation_input)

        self.layout.add_widget(Label(text="Quantité:"))
        self.quantite_input = TextInput()
        self.layout.add_widget(self.quantite_input)

        # Ajout des boutons pour valider et filtrer les données
        valider_button = Button(text="Valider")
        valider_button.bind(on_press=self.valider)
        self.layout.add_widget(valider_button)

        filtrer_button = Button(text="Filtrer")
        filtrer_button.bind(on_press=self.filtrer)
        self.layout.add_widget(filtrer_button)

        # Retourne la mise en page comme élément racine de l'application
        return self.layout

    # Fonction de validation pour enregistrer les données dans le fichier Excel
    def valider(self, instance):
        designation = self.designation_input.text
        quantite = self.quantite_input.text

        if not designation or not quantite:
            self.afficher_message("Erreur", "Veuillez remplir tous les champs.")
            return

        classeur = openpyxl.load_workbook("formulaire.xlsx")
        feuille = classeur.active
        feuille.append([designation, quantite])
        classeur.save("formulaire.xlsx")

        self.afficher_message("Succès", "Valeurs enregistrées avec succès.")
        self.designation_input.text = ""
        self.quantite_input.text = ""

    # Fonction pour filtrer les données dans le fichier Excel
    def filtrer(self, instance):
        nom_filtre = self.designation_input.text

        classeur = openpyxl.load_workbook("formulaire.xlsx")
        feuille = classeur.active

        classeur.create_sheet(nom_filtre)
        feuille_filtree = classeur[nom_filtre]

        for col in range(1, feuille.max_column + 1):
            feuille_filtree.cell(row=1, column=col, value=feuille.cell(row=1, column=col).value)

        row_filtree = 2
        for row in range(2, feuille.max_row + 1):
            if feuille.cell(row=row, column=1).value == nom_filtre:
                for col in range(1, feuille.max_column + 1):
                    feuille_filtree.cell(row=row_filtree, column=col, value=feuille.cell(row=row, column=col).value)
                row_filtree += 1

        classeur.save("formulaire.xlsx")
        self.afficher_message("Succès", "Filtrage effectué avec succès.")

    # Fonction pour afficher des messages pop-up
    def afficher_message(self, titre, message):
        popup = Popup(title=titre, content=Label(text=message), size_hint=(None, None), size=(400, 200))
        popup.open()

# Lancement de l'application Kivy
if __name__ == '__main__':
    FormulaireApp().run()
